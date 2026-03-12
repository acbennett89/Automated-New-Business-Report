from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from math import ceil
from pathlib import Path
import shutil
import tempfile
from typing import Dict, Iterable, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent if SCRIPT_DIR.name.casefold() == "scripts" else SCRIPT_DIR
WORKING_FILES_DIR = PROJECT_ROOT / "Working Files"

BIGNITION_PATH = WORKING_FILES_DIR / "Bignition_OppsByProducer.xlsx"
PRODUCTION_PATH = WORKING_FILES_DIR / "Production Report.xlsx"
OUTPUT_PATH = PROJECT_ROOT / "Consolidated New Biz Report.xlsx"
EXCLUDED_ACCOUNTS_PATH = WORKING_FILES_DIR / "Excluded Accounts.xlsx"
VELOCITY_YEAR = 2026

MONTH_NUMBERS = list(range(1, 13))
MONTH_LABELS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
PRIMARY_DEPARTMENTS = ["Commercial", "Surety", "Employee Benefits", "Personal Lines"]
SECTION_TITLE_FILL = PatternFill(fill_type="solid", fgColor="0B2A4A")
SECTION_TITLE_FONT = Font(color="FFFFFF", bold=True, size=13)
HEADER_FONT = Font(bold=True)

DEPARTMENT_SUFFIX = {
    "commercial": "01",
    "surety": "02",
    "employee benefits": "03",
}

BIGNITION_REQUIRED = (
    "Company",
    "Company Code",
    "Department",
    "Producer",
    "Potential Revenue",
    "Commission",
    "Win/Loss Date",
)
PRODUCTION_REQUIRED = ("LookupCode", "DepartmentName", "AgencyCommissionAmount", "AccountingMonth")


def normalize_header(value: object) -> str:
    return str(value).strip().casefold() if value is not None else ""


def to_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    if not text:
        return 0.0
    try:
        result = float(text)
        return -result if negative else result
    except ValueError:
        return 0.0


def to_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    formats = ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d")
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def accounting_month_to_index(value: object) -> int | None:
    if value is None:
        return None

    if isinstance(value, (int, float)):
        token = str(int(value))
    else:
        token = str(value).strip()

    if len(token) != 6 or not token.isdigit():
        return None

    year = int(token[:4])
    month = int(token[4:])
    if month < 1 or month > 12:
        return None
    return year * 12 + (month - 1)


def normalized_text(value: object) -> str:
    return str(value).strip().casefold() if value is not None else ""


def canonical_department(value: object) -> str:
    norm = normalized_text(value)
    if norm == "commercial":
        return "Commercial"
    if norm == "surety":
        return "Surety"
    if norm == "employee benefits":
        return "Employee Benefits"
    text = str(value).strip() if value is not None else ""
    return text or "Personal Lines"


def production_department(department_name: object, lookup_code: object) -> str:
    dep = normalized_text(department_name)
    if "commercial" in dep:
        return "Commercial"
    if "surety" in dep:
        return "Surety"
    if "benefit" in dep:
        return "Employee Benefits"

    lookup = str(lookup_code).strip().casefold() if lookup_code is not None else ""
    if lookup.endswith("-01"):
        return "Commercial"
    if lookup.endswith("-02"):
        return "Surety"
    if lookup.endswith("-03"):
        return "Employee Benefits"
    return "Personal Lines"


def find_header_row_and_map(ws, required_headers: Iterable[str], max_scan_rows: int = 50) -> Tuple[int, Dict[str, int]]:
    required_norm = [normalize_header(h) for h in required_headers]

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        row_values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [normalize_header(v) for v in row_values]
        index_map: Dict[str, int] = {}
        ok = True
        for required in required_norm:
            if required not in normalized:
                ok = False
                break
            index_map[required] = normalized.index(required)
        if ok:
            return row_idx, index_map

    raise ValueError(f"ERROR: Required headers not found: {list(required_headers)}")


def find_sheet_header_row_and_map(
    wb,
    required_headers: Iterable[str],
    max_scan_rows: int = 100,
):
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header_row, index_map = find_header_row_and_map(ws, required_headers, max_scan_rows=max_scan_rows)
            return ws, header_row, index_map
        except ValueError:
            continue
    raise ValueError(f"ERROR: Required headers not found in any worksheet: {list(required_headers)}")


def readable_input_path(path: Path) -> Path:
    try:
        with path.open("rb"):
            return path
    except PermissionError:
        temp_copy = Path(tempfile.gettempdir()) / f"{path.stem}_readcopy{path.suffix}"
        shutil.copy2(path, temp_copy)
        return temp_copy


def save_with_fallback(wb: Workbook, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}.new{path.suffix}")
        wb.save(fallback)
        return fallback


def load_excluded_lookup_codes(path: Path) -> set[str]:
    if not path.is_file():
        return set()

    wb = load_workbook(readable_input_path(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    header_row, index_map = find_header_row_and_map(ws, ("Account Lookup Code",))
    lookup_idx = index_map[normalize_header("Account Lookup Code")]

    result: set[str] = set()
    for row in ws.iter_rows(min_row=header_row + 1, max_col=lookup_idx + 1, values_only=True):
        raw = row[lookup_idx] if row and lookup_idx < len(row) else None
        code = normalized_text(raw)
        if code:
            result.add(code)

    wb.close()
    return result


def build_production_aggregates(
    path: Path,
    excluded_lookup_codes: set[str],
) -> tuple[dict[str, dict[str, object]], dict[str, list[float]], dict[str, list[float]]]:
    wb = load_workbook(readable_input_path(path), read_only=True, data_only=True)
    ws, header_row, index_map = find_sheet_header_row_and_map(wb, PRODUCTION_REQUIRED, max_scan_rows=200)
    lookup_idx = index_map[normalize_header("LookupCode")]
    department_idx = index_map[normalize_header("DepartmentName")]
    amount_idx = index_map[normalize_header("AgencyCommissionAmount")]
    acct_month_idx = index_map[normalize_header("AccountingMonth")]
    required_max_col = max(lookup_idx, department_idx, amount_idx, acct_month_idx) + 1

    aggregates: dict[str, dict[str, object]] = {}
    all_sales_by_department_2025: dict[str, list[float]] = defaultdict(lambda: [0.0] * 12)
    filtered_sales_by_department_2025: dict[str, list[float]] = defaultdict(lambda: [0.0] * 12)
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_col=required_max_col,
        values_only=True,
    ):
        if row is None:
            continue
        lookup_raw = row[lookup_idx] if lookup_idx < len(row) else None
        lookup_code = str(lookup_raw).strip() if lookup_raw is not None else ""
        if not lookup_code:
            continue
        lookup_norm = normalized_text(lookup_code)
        department_raw = row[department_idx] if department_idx < len(row) else None
        amount_raw = row[amount_idx] if amount_idx < len(row) else None
        amount = to_number(amount_raw)
        month_idx = accounting_month_to_index(row[acct_month_idx] if acct_month_idx < len(row) else None)

        if lookup_code not in aggregates:
            aggregates[lookup_code] = {
                "total": 0.0,
                "monthly": defaultdict(float),
            }
        aggregates[lookup_code]["total"] = float(aggregates[lookup_code]["total"]) + amount
        if month_idx is not None:
            monthly = aggregates[lookup_code]["monthly"]
            monthly[month_idx] += amount

            # Denominator: all production sales in same month last year (2025), by department.
            year = month_idx // 12
            month_pos = month_idx % 12
            if year == VELOCITY_YEAR - 1:
                dep_name = production_department(department_raw, lookup_code)
                all_sales_by_department_2025[dep_name][month_pos] += amount
                if lookup_norm not in excluded_lookup_codes:
                    filtered_sales_by_department_2025[dep_name][month_pos] += amount

    wb.close()
    return aggregates, all_sales_by_department_2025, filtered_sales_by_department_2025


def year_window_amounts(aggregate: dict[str, object] | None, win_loss_date: date | None) -> tuple[float, float]:
    if not aggregate or win_loss_date is None:
        return 0.0, 0.0

    start_idx = win_loss_date.year * 12 + (win_loss_date.month - 1)
    year1 = 0.0
    year2 = 0.0

    monthly = aggregate.get("monthly", {})
    for month_idx, amount in monthly.items():
        delta = month_idx - start_idx
        if 0 <= delta <= 11:
            year1 += amount
        elif 12 <= delta <= 23:
            year2 += amount

    return year1, year2


def month_index_from_date(value: date) -> int:
    return value.year * 12 + (value.month - 1)


def is_earlier_date(candidate: date | None, current: date | None) -> bool:
    if candidate is None:
        return False
    if current is None:
        return True
    return candidate < current


def build_bignition_aggregated_rows(ws, header_row: int, index_map: Dict[str, int]) -> tuple[list[object], list[list[object]]]:
    company_code_idx = index_map[normalize_header("Company Code")]
    department_idx = index_map[normalize_header("Department")]
    producer_idx = index_map[normalize_header("Producer")]
    potential_idx = index_map[normalize_header("Potential Revenue")]
    win_loss_idx = index_map[normalize_header("Win/Loss Date")]

    header_values = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
    groups: dict[tuple[str, str, str, str], dict[str, object]] = {}

    for row_number, row in enumerate(ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
        values_only=True,
    ), start=header_row + 1):
        row_values = list(row)
        company_code_value = row_values[company_code_idx] if company_code_idx < len(row_values) else None
        department_value = row_values[department_idx] if department_idx < len(row_values) else None
        producer_value = row_values[producer_idx] if producer_idx < len(row_values) else None
        epic_lookup = build_epic_lookup(company_code_value, department_value)

        company_code_text = str(company_code_value).strip() if company_code_value is not None else ""
        if not company_code_text:
            # Keep blank company-code rows as distinct records (no dedupe).
            key = ("__blank_company_code__", str(row_number))
        else:
            key = (
                normalized_text(epic_lookup),
                normalized_text(department_value),
                normalized_text(producer_value),
            )

        potential_value = to_number(row_values[potential_idx] if potential_idx < len(row_values) else None)
        win_loss_value = to_date(row_values[win_loss_idx] if win_loss_idx < len(row_values) else None)

        if key not in groups:
            groups[key] = {
                "row": row_values,
                "sum_potential": potential_value,
                "win_loss_date": win_loss_value,
            }
            continue

        group = groups[key]
        group["sum_potential"] = float(group["sum_potential"]) + potential_value

        current_date = group["win_loss_date"]
        if is_earlier_date(win_loss_value, current_date):
            group["row"] = row_values
            group["win_loss_date"] = win_loss_value

    aggregated_rows: list[list[object]] = []
    for group in groups.values():
        out_row = list(group["row"])
        out_row[potential_idx] = float(group["sum_potential"])
        aggregated_rows.append(out_row)

    return header_values, aggregated_rows


def build_epic_lookup(company_code: object, department: object) -> str:
    company = str(company_code).strip() if company_code is not None else ""
    if not company:
        return ""

    dept_norm = normalize_header(department)
    suffix = DEPARTMENT_SUFFIX.get(dept_norm)
    if suffix is None:
        return "ERROR"
    return f"{company}-{suffix}"


def ordered_departments(departments: Iterable[str]) -> list[str]:
    normalized = {str(d).strip(): d for d in departments if str(d).strip()}
    ordered: list[str] = []
    for dep in PRIMARY_DEPARTMENTS:
        if dep in normalized:
            ordered.append(dep)
        else:
            ordered.append(dep)
    for dep in sorted(normalized.keys()):
        if dep not in ordered:
            ordered.append(dep)
    return ordered


def auto_fit_row_heights(ws, min_row: int = 1, max_row: int | None = None, base_height: float = 15.0) -> None:
    if max_row is None:
        max_row = ws.max_row

    for row_idx in range(min_row, max_row + 1):
        max_lines = 1
        for cell in ws[row_idx]:
            if cell.value is None:
                continue

            text = str(cell.value)
            if not text:
                continue

            explicit_lines = text.count("\n") + 1

            col_letter = get_column_letter(cell.column)
            col_width = ws.column_dimensions[col_letter].width
            if col_width is None:
                col_width = 10.0

            approx_chars_per_line = max(1, int(col_width))
            wrapped_lines = ceil(len(text) / approx_chars_per_line)

            wrap_enabled = bool(cell.alignment and cell.alignment.wrap_text)
            line_count = max(explicit_lines, wrapped_lines) if wrap_enabled else explicit_lines
            if line_count > max_lines:
                max_lines = line_count

        ws.row_dimensions[row_idx].height = base_height * max_lines


def auto_fit_columns_by_content(
    ws,
    min_col: int = 1,
    max_col: int | None = None,
    min_width: float = 8.0,
    max_width: float = 36.0,
) -> None:
    if max_col is None:
        max_col = ws.max_column

    merged_ranges = list(ws.merged_cells.ranges)

    for col_idx in range(min_col, max_col + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is None:
                continue

            # Skip merged title cells so one long title does not blow out column width.
            in_merged = False
            for merged in merged_ranges:
                if cell.coordinate in merged:
                    in_merged = True
                    break
            if in_merged:
                continue

            text = str(cell.value)
            if len(text) > max_len:
                max_len = len(text)

        target_width = max(min_width, min(max_width, float(max_len + 2)))
        ws.column_dimensions[col_letter].width = target_width


def velocity_month_context(target_year: int, as_of_date: date) -> tuple[list[int], int]:
    # If running before target year, show January as current month with no full months.
    if as_of_date.year < target_year:
        return [], 0
    # If running after target year, treat December as current month and Jan-Nov as full months.
    if as_of_date.year > target_year:
        return list(range(0, 11)), 11

    current_month_pos = max(0, min(11, as_of_date.month - 1))
    full_month_positions = list(range(0, current_month_pos))
    return full_month_positions, current_month_pos


def write_compact_velocity_section(
    ws,
    start_row: int,
    title: str,
    departments: list[str],
    numerator_by_dept: dict[str, list[float]],
    denominator_by_dept: dict[str, list[float]] | None,
    value_kind: str,
    full_month_positions: list[int],
    current_month_pos: int,
) -> int:
    month_cols_start = 2
    ytd_full_col = month_cols_start + len(full_month_positions)
    spacer_col = ytd_full_col + 1
    current_month_col = spacer_col + 1
    ytd_current_col = current_month_col + 1
    end_col = ytd_current_col

    # Title row styling.
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = SECTION_TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in range(1, end_col + 1):
        ws.cell(row=start_row, column=col).fill = SECTION_TITLE_FILL

    header_row = start_row + 1
    ws.cell(row=header_row, column=1, value="Department").font = HEADER_FONT
    ws.column_dimensions[get_column_letter(1)].width = 32

    for i, month_pos in enumerate(full_month_positions):
        col = month_cols_start + i
        ws.cell(row=header_row, column=col, value=MONTH_LABELS[month_pos]).font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = 8

    ws.cell(row=header_row, column=ytd_full_col, value="YTD - Full Months").font = HEADER_FONT
    ws.column_dimensions[get_column_letter(ytd_full_col)].width = 14

    ws.cell(row=header_row, column=spacer_col, value="").font = HEADER_FONT
    ws.column_dimensions[get_column_letter(spacer_col)].width = 2.5

    ws.cell(row=header_row, column=current_month_col, value=MONTH_LABELS[current_month_pos]).font = HEADER_FONT
    ws.column_dimensions[get_column_letter(current_month_col)].width = 8

    ws.cell(row=header_row, column=ytd_current_col, value="YTD - Current Month").font = HEADER_FONT
    ws.column_dimensions[get_column_letter(ytd_current_col)].width = 16

    totals_num = [0.0] * 12
    totals_den = [0.0] * 12

    row = start_row + 2
    for dept in departments:
        ws.cell(row=row, column=1, value=dept)
        num_values = numerator_by_dept.get(dept, [0.0] * 12)
        den_values = denominator_by_dept.get(dept, [0.0] * 12) if denominator_by_dept is not None else None

        full_num_sum = 0.0
        full_den_sum = 0.0

        for i, month_pos in enumerate(full_month_positions):
            col = month_cols_start + i
            num = num_values[month_pos]
            den = den_values[month_pos] if den_values is not None else 1.0
            totals_num[month_pos] += num
            if den_values is not None:
                totals_den[month_pos] += den
            full_num_sum += num
            if den_values is not None:
                full_den_sum += den

            value = round(num) if den_values is None else ((num / den) if den else 0.0)
            cell = ws.cell(row=row, column=col, value=value)
            cell.number_format = "$#,##0" if value_kind == "money" else "0.00%"

        current_num = num_values[current_month_pos]
        current_den = den_values[current_month_pos] if den_values is not None else 1.0
        totals_num[current_month_pos] += current_num
        if den_values is not None:
            totals_den[current_month_pos] += current_den

        ytd_current_num = full_num_sum + current_num
        ytd_current_den = full_den_sum + (current_den if den_values is not None else 0.0)

        ytd_full_value = round(full_num_sum) if den_values is None else ((full_num_sum / full_den_sum) if full_den_sum else 0.0)
        current_value = round(current_num) if den_values is None else ((current_num / current_den) if current_den else 0.0)
        ytd_current_value = (
            round(ytd_current_num)
            if den_values is None
            else ((ytd_current_num / ytd_current_den) if ytd_current_den else 0.0)
        )

        ytd_full_cell = ws.cell(row=row, column=ytd_full_col, value=ytd_full_value)
        current_cell = ws.cell(row=row, column=current_month_col, value=current_value)
        ytd_current_cell = ws.cell(row=row, column=ytd_current_col, value=ytd_current_value)
        for c in (ytd_full_cell, current_cell, ytd_current_cell):
            c.number_format = "$#,##0" if value_kind == "money" else "0.00%"

        row += 1

    total_row = row
    total_label = ws.cell(row=total_row, column=1, value="Total")
    total_label.font = HEADER_FONT

    full_total_num = sum(totals_num[pos] for pos in full_month_positions)
    full_total_den = sum(totals_den[pos] for pos in full_month_positions)
    current_total_num = totals_num[current_month_pos]
    current_total_den = totals_den[current_month_pos]
    ytd_current_total_num = full_total_num + current_total_num
    ytd_current_total_den = full_total_den + current_total_den

    for i, month_pos in enumerate(full_month_positions):
        col = month_cols_start + i
        num = totals_num[month_pos]
        den = totals_den[month_pos]
        value = round(num) if denominator_by_dept is None else ((num / den) if den else 0.0)
        cell = ws.cell(row=total_row, column=col, value=value)
        cell.font = HEADER_FONT
        cell.number_format = "$#,##0" if value_kind == "money" else "0.00%"

    total_ytd_full = (
        round(full_total_num)
        if denominator_by_dept is None
        else ((full_total_num / full_total_den) if full_total_den else 0.0)
    )
    total_current = (
        round(current_total_num)
        if denominator_by_dept is None
        else ((current_total_num / current_total_den) if current_total_den else 0.0)
    )
    total_ytd_current = (
        round(ytd_current_total_num)
        if denominator_by_dept is None
        else ((ytd_current_total_num / ytd_current_total_den) if ytd_current_total_den else 0.0)
    )

    total_ytd_full_cell = ws.cell(row=total_row, column=ytd_full_col, value=total_ytd_full)
    total_current_cell = ws.cell(row=total_row, column=current_month_col, value=total_current)
    total_ytd_current_cell = ws.cell(row=total_row, column=ytd_current_col, value=total_ytd_current)
    for c in (total_ytd_full_cell, total_current_cell, total_ytd_current_cell):
        c.font = HEADER_FONT
        c.number_format = "$#,##0" if value_kind == "money" else "0.00%"

    return total_row + 2


def write_sales_velocity_sheet(
    wb_out: Workbook,
    sheet_name: str,
    departments: list[str],
    newbiz_2026_by_dept: dict[str, list[float]],
    all_sales_by_department_2025: dict[str, list[float]],
) -> None:
    ws_velocity = wb_out.create_sheet(sheet_name)
    full_month_positions, current_month_pos = velocity_month_context(VELOCITY_YEAR, date.today())

    next_row = write_compact_velocity_section(
        ws_velocity,
        start_row=1,
        title=f"Sales Velocity ({VELOCITY_YEAR} New Business / {VELOCITY_YEAR - 1} Same Month Sales)",
        departments=departments,
        numerator_by_dept=newbiz_2026_by_dept,
        denominator_by_dept=all_sales_by_department_2025,
        value_kind="velocity",
        full_month_positions=full_month_positions,
        current_month_pos=current_month_pos,
    )
    next_row = write_compact_velocity_section(
        ws_velocity,
        start_row=next_row,
        title=f"Aggregated New Business Production ({VELOCITY_YEAR})",
        departments=departments,
        numerator_by_dept=newbiz_2026_by_dept,
        denominator_by_dept=None,
        value_kind="money",
        full_month_positions=full_month_positions,
        current_month_pos=current_month_pos,
    )
    write_compact_velocity_section(
        ws_velocity,
        start_row=next_row,
        title=f"Aggregated Same Month Prior Year Sales ({VELOCITY_YEAR - 1})",
        departments=departments,
        numerator_by_dept=all_sales_by_department_2025,
        denominator_by_dept=None,
        value_kind="money",
        full_month_positions=full_month_positions,
        current_month_pos=current_month_pos,
    )
    auto_fit_row_heights(ws_velocity)
    auto_fit_columns_by_content(ws_velocity)


def consolidate() -> None:
    if not BIGNITION_PATH.is_file():
        raise FileNotFoundError(f"ERROR: Missing file: {BIGNITION_PATH}")
    if not PRODUCTION_PATH.is_file():
        raise FileNotFoundError(f"ERROR: Missing file: {PRODUCTION_PATH}")

    print(f"Reading Bignition file: {BIGNITION_PATH}")
    print(f"Reading Production file: {PRODUCTION_PATH}")

    excluded_lookup_codes = load_excluded_lookup_codes(EXCLUDED_ACCOUNTS_PATH)
    print(f"Loaded excluded lookup codes: {len(excluded_lookup_codes)}")

    production_aggregates, all_sales_by_department_2025, filtered_sales_by_department_2025 = build_production_aggregates(
        PRODUCTION_PATH,
        excluded_lookup_codes,
    )
    print(f"Loaded production totals for {len(production_aggregates)} lookup codes.")

    wb_in = load_workbook(readable_input_path(BIGNITION_PATH), read_only=True, data_only=True)
    ws_in, header_row, index_map = find_sheet_header_row_and_map(wb_in, BIGNITION_REQUIRED, max_scan_rows=200)
    department_idx = index_map[normalize_header("Department")] + 1
    company_idx = index_map[normalize_header("Company Code")] + 1
    commission_idx = index_map[normalize_header("Commission")] + 1
    win_loss_idx = index_map[normalize_header("Win/Loss Date")] + 1

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Consolidated New Biz Report"

    # Start output at the actual header row (drops top label rows), then aggregate duplicates.
    header_values, aggregated_rows = build_bignition_aggregated_rows(ws_in, header_row, index_map)
    ws_out.append(header_values)
    for row in aggregated_rows:
        ws_out.append(row)

    print(f"Aggregated Bignition rows: {len(aggregated_rows)}")

    epic_lookup_col = ws_in.max_column + 1
    total_billed_col = ws_in.max_column + 2
    amount_year_1_col = ws_in.max_column + 3
    amount_year_2_col = ws_in.max_column + 4

    ws_out.cell(row=1, column=epic_lookup_col, value="EPIC Lookup Code")
    ws_out.cell(row=1, column=total_billed_col, value="Total Billed to date")
    ws_out.cell(row=1, column=amount_year_1_col, value="Amount Billed Year 1")
    ws_out.cell(row=1, column=amount_year_2_col, value="Amount Billed Year 2")

    # For Sales Velocity tab.
    newbiz_2026_by_dept: dict[str, list[float]] = defaultdict(lambda: [0.0] * 12)
    filtered_newbiz_2026_by_dept: dict[str, list[float]] = defaultdict(lambda: [0.0] * 12)

    error_count = 0
    for row_idx in range(2, ws_out.max_row + 1):
        company_code = ws_out.cell(row=row_idx, column=company_idx).value
        department = ws_out.cell(row=row_idx, column=department_idx).value
        department_name = canonical_department(department)
        commission_factor = to_number(ws_out.cell(row=row_idx, column=commission_idx).value)
        win_loss_date = to_date(ws_out.cell(row=row_idx, column=win_loss_idx).value)

        epic_lookup = build_epic_lookup(company_code, department)
        ws_out.cell(row=row_idx, column=epic_lookup_col, value=epic_lookup)

        if epic_lookup == "ERROR":
            error_count += 1
            ws_out.cell(row=row_idx, column=total_billed_col, value=0.0)
            ws_out.cell(row=row_idx, column=amount_year_1_col, value=0.0)
            ws_out.cell(row=row_idx, column=amount_year_2_col, value=0.0)
            continue

        aggregate = production_aggregates.get(epic_lookup) if epic_lookup else None
        total_raw = float(aggregate["total"]) if aggregate else 0.0
        year1_raw, year2_raw = year_window_amounts(aggregate, win_loss_date)

        ws_out.cell(row=row_idx, column=total_billed_col, value=total_raw * commission_factor)
        ws_out.cell(row=row_idx, column=amount_year_1_col, value=year1_raw * commission_factor)
        ws_out.cell(row=row_idx, column=amount_year_2_col, value=year2_raw * commission_factor)

        # Build monthly Sales Velocity metrics for 2026.
        if aggregate and win_loss_date is not None:
            start_idx = month_index_from_date(win_loss_date)
            monthly = aggregate.get("monthly", {})
            is_excluded = normalized_text(epic_lookup) in excluded_lookup_codes
            for month_num in MONTH_NUMBERS:
                month_pos = month_num - 1
                current_idx = VELOCITY_YEAR * 12 + month_pos
                # New business: account within first year from win/loss month.
                if start_idx <= current_idx <= start_idx + 11:
                    current_amount = monthly.get(current_idx, 0.0) * commission_factor
                    newbiz_2026_by_dept[department_name][month_pos] += current_amount
                    if not is_excluded:
                        filtered_newbiz_2026_by_dept[department_name][month_pos] += current_amount

    # Format the full output as an Excel table.
    last_col_letter = get_column_letter(ws_out.max_column)
    table_ref = f"A1:{last_col_letter}{ws_out.max_row}"
    output_table = Table(displayName="ConsolidatedNewBizTable", ref=table_ref)
    output_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_out.add_table(output_table)

    all_depts = set(newbiz_2026_by_dept.keys()) | set(all_sales_by_department_2025.keys())
    departments = ordered_departments(all_depts)
    write_sales_velocity_sheet(
        wb_out=wb_out,
        sheet_name="Sales Velocity 2026",
        departments=departments,
        newbiz_2026_by_dept=newbiz_2026_by_dept,
        all_sales_by_department_2025=all_sales_by_department_2025,
    )

    filtered_depts = set(filtered_newbiz_2026_by_dept.keys()) | set(filtered_sales_by_department_2025.keys())
    filtered_departments = ordered_departments(filtered_depts)
    write_sales_velocity_sheet(
        wb_out=wb_out,
        sheet_name="Sales Velocity 2026 Excluded",
        departments=filtered_departments,
        newbiz_2026_by_dept=filtered_newbiz_2026_by_dept,
        all_sales_by_department_2025=filtered_sales_by_department_2025,
    )

    saved_path = save_with_fallback(wb_out, OUTPUT_PATH)
    wb_in.close()

    print(f"Wrote output file: {saved_path}")
    if error_count:
        print(f"ERROR: {error_count} row(s) had an unmatched Department for EPIC Lookup Code mapping.")
    else:
        print("Completed with no Department mapping errors.")


def main() -> int:
    try:
        consolidate()
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
