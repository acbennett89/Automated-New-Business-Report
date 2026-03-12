from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import shutil
import tempfile
from typing import Dict, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Font


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent if SCRIPT_DIR.name.casefold() == "scripts" else SCRIPT_DIR

CONSOLIDATED_PATH = PROJECT_ROOT / "Consolidated New Biz Report.xlsx"
TARGET_DEPARTMENTS = ("Commercial", "Surety", "Employee Benefits")
REQUIRED_HEADERS = (
    "Company",
    "Company Code",
    "Status",
    "Producer",
    "Department",
    "Win/Loss Date",
    "Potential Revenue",
    "Total Billed to date",
    "Amount Billed Year 1",
)


def normalize_header(value: object) -> str:
    return str(value).strip().casefold() if value is not None else ""


def normalized_text(value: object) -> str:
    return str(value).strip().casefold() if value is not None else ""


def to_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    if not text:
        return 0.0
    try:
        result = float(text)
        return -result if negative else result
    except ValueError:
        return 0.0


def to_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def canonical_department(value: object) -> str:
    norm = normalized_text(value)
    if norm == "commercial":
        return "Commercial"
    if norm == "surety":
        return "Surety"
    if norm == "employee benefits":
        return "Employee Benefits"
    return ""


def status_is_customer(value: object) -> bool:
    return "customer" in normalized_text(value)


def find_header_row_and_map(ws, required_headers: Iterable[str], max_scan_rows: int = 50) -> tuple[int, Dict[str, int]]:
    required_norm = [normalize_header(h) for h in required_headers]

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [normalize_header(v) for v in values]
        idx_map: Dict[str, int] = {}
        ok = True
        for needed in required_norm:
            if needed not in normalized:
                ok = False
                break
            idx_map[needed] = normalized.index(needed)
        if ok:
            return row_idx, idx_map

    raise ValueError(f"ERROR: Required headers not found: {list(required_headers)}")


def readable_input_path(path: Path) -> Path:
    try:
        with path.open("rb"):
            return path
    except PermissionError:
        temp_copy = Path(tempfile.gettempdir()) / f"{path.stem}_readcopy{path.suffix}"
        shutil.copy2(path, temp_copy)
        return temp_copy


def save_with_fallback(wb, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}.new{path.suffix}")
        wb.save(fallback)
        return fallback


def prorata_multiplier(run_date: date, win_date: date) -> float:
    days = (run_date - win_date).days
    raw = days / 365.0
    if raw < 0:
        return 0.0
    if raw > 1:
        return 1.0
    return raw


def load_customer_rows(ws, header_row: int, index_map: Dict[str, int]) -> list[dict[str, object]]:
    company_idx = index_map[normalize_header("Company")]
    status_idx = index_map[normalize_header("Status")]
    producer_idx = index_map[normalize_header("Producer")]
    department_idx = index_map[normalize_header("Department")]
    win_loss_idx = index_map[normalize_header("Win/Loss Date")]
    potential_idx = index_map[normalize_header("Potential Revenue")]
    total_billed_idx = index_map[normalize_header("Total Billed to date")]
    amount_year_1_idx = index_map[normalize_header("Amount Billed Year 1")]
    max_col = max(
        company_idx,
        status_idx,
        producer_idx,
        department_idx,
        win_loss_idx,
        potential_idx,
        total_billed_idx,
        amount_year_1_idx,
    ) + 1

    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True):
        if row is None:
            continue
        status = row[status_idx] if status_idx < len(row) else None
        if not status_is_customer(status):
            continue

        win_date = to_date(row[win_loss_idx] if win_loss_idx < len(row) else None)
        if win_date is None:
            continue

        department = canonical_department(row[department_idx] if department_idx < len(row) else None)
        if department not in TARGET_DEPARTMENTS:
            continue

        rows.append(
            {
                "company": str(row[company_idx]).strip() if company_idx < len(row) and row[company_idx] is not None else "",
                "producer": str(row[producer_idx]).strip() if producer_idx < len(row) and row[producer_idx] is not None else "",
                "department": department,
                "win_date": win_date,
                "potential_revenue": to_number(row[potential_idx] if potential_idx < len(row) else None),
                "total_billed_to_date": to_number(row[total_billed_idx] if total_billed_idx < len(row) else None),
                "amount_billed_year_1": to_number(row[amount_year_1_idx] if amount_year_1_idx < len(row) else None),
            }
        )
    return rows


def load_no_key_rows(ws, header_row: int, index_map: Dict[str, int], run_date: date) -> list[dict[str, object]]:
    company_idx = index_map[normalize_header("Company")]
    company_code_idx = index_map[normalize_header("Company Code")]
    status_idx = index_map[normalize_header("Status")]
    department_idx = index_map[normalize_header("Department")]
    win_loss_idx = index_map[normalize_header("Win/Loss Date")]
    max_col = max(company_idx, company_code_idx, status_idx, department_idx, win_loss_idx) + 1

    allowed_years = {run_date.year - 1, run_date.year}
    rows: list[dict[str, object]] = []
    for row in ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True):
        if row is None:
            continue

        company_code = row[company_code_idx] if company_code_idx < len(row) else None
        company_code_text = str(company_code).strip() if company_code is not None else ""
        if company_code_text:
            continue

        status = row[status_idx] if status_idx < len(row) else None
        if not status_is_customer(status):
            continue

        win_date = to_date(row[win_loss_idx] if win_loss_idx < len(row) else None)
        if win_date is None or win_date.year not in allowed_years:
            continue

        rows.append(
            {
                "company": str(row[company_idx]).strip() if company_idx < len(row) and row[company_idx] is not None else "",
                "department": str(row[department_idx]).strip() if department_idx < len(row) and row[department_idx] is not None else "",
                "win_date": win_date,
            }
        )

    rows.sort(key=lambda r: (str(r["company"]).casefold(), str(r["department"]).casefold(), r["win_date"]))
    return rows


def auto_fit_columns(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = ws.cell(row=1, column=col).column_letter
        max_len = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            length = len(str(value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def write_year_tab(wb, year: int, rows: list[dict[str, object]], run_date: date) -> None:
    sheet_name = f"{year} New Biz"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    row_idx = 1
    ws.cell(row=row_idx, column=1, value=sheet_name).font = Font(bold=True, size=14)
    row_idx += 2

    headers = [
        "Producer",
        "Account",
        "Win/Loss Date",
        "Potential Revenue",
        "Total Billed to Date",
        "Amount Billed Year 1",
        "Prorated Expected Billings",
    ]

    for department in TARGET_DEPARTMENTS:
        ws.cell(row=row_idx, column=1, value=department).font = Font(bold=True)
        row_idx += 1

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=header).font = Font(bold=True)
        row_idx += 1

        dept_rows = [r for r in rows if r["department"] == department and r["win_date"].year == year]
        dept_rows.sort(key=lambda r: (str(r["producer"]).casefold(), str(r["company"]).casefold(), r["win_date"]))

        if not dept_rows:
            ws.cell(row=row_idx, column=1, value="No matching customer accounts.")
            row_idx += 2
            continue

        total_potential = 0.0
        total_billed = 0.0
        total_year_1 = 0.0
        total_prorated = 0.0

        for item in dept_rows:
            multiplier = prorata_multiplier(run_date, item["win_date"])
            prorated_expected = float(item["potential_revenue"]) * multiplier

            ws.cell(row=row_idx, column=1, value=item["producer"])
            ws.cell(row=row_idx, column=2, value=item["company"])
            win_cell = ws.cell(row=row_idx, column=3, value=item["win_date"])
            win_cell.number_format = "mm/dd/yyyy"

            pot_cell = ws.cell(row=row_idx, column=4, value=float(item["potential_revenue"]))
            billed_cell = ws.cell(row=row_idx, column=5, value=float(item["total_billed_to_date"]))
            year1_cell = ws.cell(row=row_idx, column=6, value=float(item["amount_billed_year_1"]))
            prorated_cell = ws.cell(row=row_idx, column=7, value=prorated_expected)

            pot_cell.number_format = "$#,##0"
            billed_cell.number_format = "$#,##0"
            year1_cell.number_format = "$#,##0"
            prorated_cell.number_format = "$#,##0"

            total_potential += float(item["potential_revenue"])
            total_billed += float(item["total_billed_to_date"])
            total_year_1 += float(item["amount_billed_year_1"])
            total_prorated += prorated_expected
            row_idx += 1

        ws.cell(row=row_idx, column=1, value="Department Total").font = Font(bold=True)
        t_pot = ws.cell(row=row_idx, column=4, value=total_potential)
        t_billed = ws.cell(row=row_idx, column=5, value=total_billed)
        t_year1 = ws.cell(row=row_idx, column=6, value=total_year_1)
        t_prorated = ws.cell(row=row_idx, column=7, value=total_prorated)
        t_pot.font = Font(bold=True)
        t_billed.font = Font(bold=True)
        t_year1.font = Font(bold=True)
        t_prorated.font = Font(bold=True)
        t_pot.number_format = "$#,##0"
        t_billed.number_format = "$#,##0"
        t_year1.number_format = "$#,##0"
        t_prorated.number_format = "$#,##0"

        row_idx += 2

    auto_fit_columns(ws, max_col=7)


def write_no_key_tab(wb, rows: list[dict[str, object]]) -> None:
    sheet_name = "New Biz No Key"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    ws.cell(row=1, column=1, value=sheet_name).font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="Company Name").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Department").font = Font(bold=True)

    row_idx = 4
    if not rows:
        ws.cell(row=row_idx, column=1, value="No matching accounts.")
        auto_fit_columns(ws, max_col=2)
        return

    for item in rows:
        ws.cell(row=row_idx, column=1, value=item["company"])
        ws.cell(row=row_idx, column=2, value=item["department"])
        row_idx += 1

    auto_fit_columns(ws, max_col=2)


def run() -> None:
    if not CONSOLIDATED_PATH.is_file():
        raise FileNotFoundError(f"ERROR: Missing file: {CONSOLIDATED_PATH}")

    run_date = date.today()
    previous_year = run_date.year - 1
    current_year = run_date.year

    print(f"Reading consolidated workbook: {CONSOLIDATED_PATH}")
    source_path = readable_input_path(CONSOLIDATED_PATH)
    wb = load_workbook(source_path)

    ws_source = wb["Consolidated New Biz Report"] if "Consolidated New Biz Report" in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row, index_map = find_header_row_and_map(ws_source, REQUIRED_HEADERS, max_scan_rows=50)
    rows = load_customer_rows(ws_source, header_row, index_map)
    no_key_rows = load_no_key_rows(ws_source, header_row, index_map, run_date)
    print(f"Loaded customer rows for target departments: {len(rows)}")
    print(f"Loaded customer rows for 'New Biz No Key': {len(no_key_rows)}")

    write_year_tab(wb, previous_year, rows, run_date)
    write_year_tab(wb, current_year, rows, run_date)
    write_no_key_tab(wb, no_key_rows)

    saved_path = save_with_fallback(wb, CONSOLIDATED_PATH)
    wb.close()
    print(f"Wrote tabs: '{previous_year} New Biz', '{current_year} New Biz', and 'New Biz No Key'")
    print(f"Saved workbook: {saved_path}")


def main() -> int:
    try:
        run()
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
