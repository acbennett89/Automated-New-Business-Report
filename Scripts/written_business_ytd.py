from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
import shutil
import tempfile
from typing import Dict, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Font


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent if SCRIPT_DIR.name.casefold() == "scripts" else SCRIPT_DIR
CONSOLIDATED_PATH = PROJECT_ROOT / "Consolidated New Biz Report.xlsx"
TARGET_SHEET = "Consolidated New Biz Report"
OUTPUT_SHEET = "Written Business YTD vs PYTD"

PRIMARY_DEPARTMENTS = ["Commercial", "Surety", "Employee Benefits", "Personal Lines"]
REQUIRED_HEADERS = ("Producer", "Department", "Status", "Win/Loss Date", "Potential Revenue")


def normalize_header(value: object) -> str:
    return str(value).strip().casefold() if value is not None else ""


def normalized_text(value: object) -> str:
    return str(value).strip().casefold() if value is not None else ""


def to_number(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "").strip()
    if not text:
        return 0.0
    try:
        result = float(text)
        return -result if negative else result
    except ValueError:
        return 0.0


def to_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def canonical_department(value: object) -> str:
    norm = normalized_text(value)
    if norm == "commercial":
        return "Commercial"
    if norm == "surety":
        return "Surety"
    if norm == "employee benefits":
        return "Employee Benefits"
    text = str(value).strip() if value is not None else ""
    return text or "Personal Lines"


def status_is_customer(value: object) -> bool:
    return "customer" in normalized_text(value)


def find_header_row_and_map(ws, required_headers: Iterable[str], max_scan_rows: int = 50) -> tuple[int, Dict[str, int]]:
    required_norm = [normalize_header(h) for h in required_headers]

    for row_idx in range(1, min(ws.max_row, max_scan_rows) + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [normalize_header(v) for v in values]
        idx_map: Dict[str, int] = {}
        ok = True
        for needed in required_norm:
            if needed not in normalized:
                ok = False
                break
            idx_map[needed] = normalized.index(needed)
        if ok:
            return row_idx, idx_map

    raise ValueError(f"ERROR: Required headers not found: {list(required_headers)}")


def readable_input_path(path: Path) -> Path:
    try:
        with path.open("rb"):
            return path
    except PermissionError:
        temp_copy = Path(tempfile.gettempdir()) / f"{path.stem}_readcopy{path.suffix}"
        shutil.copy2(path, temp_copy)
        return temp_copy


def save_with_fallback(wb, path: Path) -> Path:
    try:
        wb.save(path)
        return path
    except PermissionError:
        fallback = path.with_name(f"{path.stem}.new{path.suffix}")
        wb.save(fallback)
        return fallback


def ordered_departments(departments: Iterable[str]) -> list[str]:
    normalized = {str(d).strip(): d for d in departments if str(d).strip()}
    ordered: list[str] = []
    for dep in PRIMARY_DEPARTMENTS:
        if dep in normalized:
            ordered.append(dep)
    for dep in sorted(normalized.keys()):
        if dep not in ordered:
            ordered.append(dep)
    return ordered


def totals_by_department(producer_map: dict[str, dict[str, float]], departments: list[str]) -> dict[str, float]:
    totals = {dep: 0.0 for dep in departments}
    for per_producer in producer_map.values():
        for dep, amount in per_producer.items():
            if dep not in totals:
                totals[dep] = 0.0
            totals[dep] += float(amount)
    return totals


def safe_pct(numerator: float, denominator: float) -> float:
    return (numerator / denominator) if denominator else 0.0


def write_summary_visual(
    ws,
    start_row: int,
    this_year: int,
    prior_year: int,
    departments: list[str],
    this_totals: dict[str, float],
    prior_totals: dict[str, float],
) -> int:
    ws.cell(row=start_row, column=1, value=f"Written Business Summary ({this_year} YTD vs {prior_year} PYTD)").font = Font(
        bold=True, size=14
    )

    overall_this = sum(this_totals.get(dep, 0.0) for dep in departments)
    overall_prior = sum(prior_totals.get(dep, 0.0) for dep in departments)
    overall_delta = overall_this - overall_prior
    overall_pct = safe_pct(overall_delta, overall_prior)

    # KPI block
    ws.cell(row=start_row + 2, column=1, value="Metric").font = Font(bold=True)
    ws.cell(row=start_row + 2, column=2, value="Value").font = Font(bold=True)

    ws.cell(row=start_row + 3, column=1, value=f"{this_year} YTD Total")
    c = ws.cell(row=start_row + 3, column=2, value=overall_this)
    c.number_format = "$#,##0"

    ws.cell(row=start_row + 4, column=1, value=f"{prior_year} PYTD Total")
    c = ws.cell(row=start_row + 4, column=2, value=overall_prior)
    c.number_format = "$#,##0"

    ws.cell(row=start_row + 5, column=1, value="Delta")
    c = ws.cell(row=start_row + 5, column=2, value=overall_delta)
    c.number_format = "$#,##0"

    ws.cell(row=start_row + 6, column=1, value="% Change")
    c = ws.cell(row=start_row + 6, column=2, value=overall_pct)
    c.number_format = "0.00%"

    # Department breakdown block
    dep_start_col = 4
    ws.cell(row=start_row + 2, column=dep_start_col, value="Department").font = Font(bold=True)
    ws.cell(row=start_row + 2, column=dep_start_col + 1, value=str(this_year)).font = Font(bold=True)
    ws.cell(row=start_row + 2, column=dep_start_col + 2, value=str(prior_year)).font = Font(bold=True)
    ws.cell(row=start_row + 2, column=dep_start_col + 3, value="Delta").font = Font(bold=True)
    ws.cell(row=start_row + 2, column=dep_start_col + 4, value="% Change").font = Font(bold=True)

    row = start_row + 3
    for dep in departments:
        cur = float(this_totals.get(dep, 0.0))
        prv = float(prior_totals.get(dep, 0.0))
        delta = cur - prv
        pct = safe_pct(delta, prv)

        ws.cell(row=row, column=dep_start_col, value=dep)
        c = ws.cell(row=row, column=dep_start_col + 1, value=cur)
        c.number_format = "$#,##0"
        c = ws.cell(row=row, column=dep_start_col + 2, value=prv)
        c.number_format = "$#,##0"
        c = ws.cell(row=row, column=dep_start_col + 3, value=delta)
        c.number_format = "$#,##0"
        c = ws.cell(row=row, column=dep_start_col + 4, value=pct)
        c.number_format = "0.00%"
        row += 1

    ws.cell(row=row, column=dep_start_col, value="Total").font = Font(bold=True)
    c = ws.cell(row=row, column=dep_start_col + 1, value=overall_this)
    c.font = Font(bold=True)
    c.number_format = "$#,##0"
    c = ws.cell(row=row, column=dep_start_col + 2, value=overall_prior)
    c.font = Font(bold=True)
    c.number_format = "$#,##0"
    c = ws.cell(row=row, column=dep_start_col + 3, value=overall_delta)
    c.font = Font(bold=True)
    c.number_format = "$#,##0"
    c = ws.cell(row=row, column=dep_start_col + 4, value=overall_pct)
    c.font = Font(bold=True)
    c.number_format = "0.00%"

    return row + 2


def aggregate_written_business(ws, header_row: int, index_map: Dict[str, int], run_date: date):
    producer_idx = index_map[normalize_header("Producer")]
    department_idx = index_map[normalize_header("Department")]
    status_idx = index_map[normalize_header("Status")]
    win_loss_idx = index_map[normalize_header("Win/Loss Date")]
    potential_idx = index_map[normalize_header("Potential Revenue")]
    max_col = max(producer_idx, department_idx, status_idx, win_loss_idx, potential_idx) + 1

    this_year = run_date.year
    prior_year = run_date.year - 1
    pytd_end = date(prior_year, run_date.month, run_date.day)

    this_ytd: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    pytd: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    departments: set[str] = set()

    for row in ws.iter_rows(min_row=header_row + 1, max_col=max_col, values_only=True):
        if row is None:
            continue
        status = row[status_idx] if status_idx < len(row) else None
        if not status_is_customer(status):
            continue

        win_date = to_date(row[win_loss_idx] if win_loss_idx < len(row) else None)
        if win_date is None:
            continue

        department = canonical_department(row[department_idx] if department_idx < len(row) else None)
        producer = str(row[producer_idx]).strip() if producer_idx < len(row) and row[producer_idx] is not None else "Unassigned"
        amount = to_number(row[potential_idx] if potential_idx < len(row) else None)

        if win_date.year == this_year and win_date <= run_date:
            this_ytd[producer][department] += amount
            departments.add(department)
        elif win_date.year == prior_year and win_date <= pytd_end:
            pytd[producer][department] += amount
            departments.add(department)

    return this_ytd, pytd, ordered_departments(departments), this_year, prior_year


def write_block(ws, start_row: int, title: str, producer_map: dict[str, dict[str, float]], departments: list[str]) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    ws.cell(row=start_row + 1, column=1, value="Producer").font = Font(bold=True)
    for i, dep in enumerate(departments, start=2):
        ws.cell(row=start_row + 1, column=i, value=dep).font = Font(bold=True)
    total_col = len(departments) + 2
    ws.cell(row=start_row + 1, column=total_col, value="Total").font = Font(bold=True)

    row_idx = start_row + 2
    producers = sorted(producer_map.keys(), key=lambda x: x.casefold())
    col_totals = [0.0] * len(departments)

    if not producers:
        ws.cell(row=row_idx, column=1, value="No matching rows.")
        return row_idx + 2

    for producer in producers:
        ws.cell(row=row_idx, column=1, value=producer)
        row_total = 0.0
        for i, dep in enumerate(departments):
            amount = float(producer_map.get(producer, {}).get(dep, 0.0))
            col_totals[i] += amount
            row_total += amount
            c = ws.cell(row=row_idx, column=i + 2, value=amount)
            c.number_format = "$#,##0"
        t = ws.cell(row=row_idx, column=total_col, value=row_total)
        t.number_format = "$#,##0"
        row_idx += 1

    ws.cell(row=row_idx, column=1, value="Total").font = Font(bold=True)
    grand_total = 0.0
    for i, amount in enumerate(col_totals):
        grand_total += amount
        c = ws.cell(row=row_idx, column=i + 2, value=amount)
        c.font = Font(bold=True)
        c.number_format = "$#,##0"
    gt = ws.cell(row=row_idx, column=total_col, value=grand_total)
    gt.font = Font(bold=True)
    gt.number_format = "$#,##0"
    return row_idx + 2


def auto_fit_columns(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        letter = ws.cell(row=1, column=col).column_letter
        max_len = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def run() -> None:
    if not CONSOLIDATED_PATH.is_file():
        raise FileNotFoundError(f"ERROR: Missing file: {CONSOLIDATED_PATH}")

    run_date = date.today()
    print(f"Reading workbook: {CONSOLIDATED_PATH}")
    source_path = readable_input_path(CONSOLIDATED_PATH)
    wb = load_workbook(source_path)
    ws_source = wb[TARGET_SHEET] if TARGET_SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row, index_map = find_header_row_and_map(ws_source, REQUIRED_HEADERS, max_scan_rows=50)

    this_ytd, pytd, departments, this_year, prior_year = aggregate_written_business(
        ws_source, header_row, index_map, run_date
    )
    print(f"Departments in scope: {', '.join(departments) if departments else 'None'}")
    this_totals = totals_by_department(this_ytd, departments)
    prior_totals = totals_by_department(pytd, departments)

    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]
    ws_out = wb.create_sheet(OUTPUT_SHEET)

    next_row = write_summary_visual(
        ws_out,
        start_row=1,
        this_year=this_year,
        prior_year=prior_year,
        departments=departments,
        this_totals=this_totals,
        prior_totals=prior_totals,
    )

    next_row = write_block(
        ws_out,
        start_row=next_row,
        title=f"Written Business YTD {this_year} (Potential Revenue)",
        producer_map=this_ytd,
        departments=departments,
    )
    write_block(
        ws_out,
        start_row=next_row,
        title=f"Written Business PYTD {prior_year} (Potential Revenue)",
        producer_map=pytd,
        departments=departments,
    )
    auto_fit_columns(ws_out, max_col=max(len(departments) + 2, 8))

    saved_path = save_with_fallback(wb, CONSOLIDATED_PATH)
    wb.close()
    print(f"Wrote tab: {OUTPUT_SHEET}")
    print(f"Saved workbook: {saved_path}")


def main() -> int:
    try:
        run()
        return 0
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
